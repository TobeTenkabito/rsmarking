from __future__ import annotations

import base64
import csv
import io
import json
import os
import re
import uuid
from pathlib import Path
from typing import Any


BASE_DIR = Path(__file__).resolve().parents[2]
_configured_artifact_dir = Path(os.getenv("AI_ARTIFACT_DIR") or "storage/ai_artifacts")
ARTIFACT_DIR = (
    _configured_artifact_dir
    if _configured_artifact_dir.is_absolute()
    else BASE_DIR / _configured_artifact_dir
).resolve()
MAX_ARTIFACT_BYTES = 20 * 1024 * 1024
MAX_TEXT_CHARS = 1_000_000

_FORMATS: dict[str, tuple[str, str, str]] = {
    "txt": (".txt", "text/plain; charset=utf-8", "file"),
    "md": (".md", "text/markdown; charset=utf-8", "file"),
    "html": (".html", "text/html; charset=utf-8", "file"),
    "json": (".json", "application/json", "file"),
    "svg": (".svg", "image/svg+xml", "image"),
    "csv": (".csv", "text/csv; charset=utf-8", "table"),
    "xlsx": (
        ".xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "table",
    ),
}
_IMAGE_MIME_TYPES = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "image/gif": ".gif",
}
_ARTIFACT_ID = re.compile(r"^[a-f0-9]{32}$")
_UNSAFE_SVG = re.compile(
    r"<(?:script|foreignObject|iframe|object|embed)\b|\bon\w+\s*=|"
    r"(?:href|xlink:href)\s*=\s*['\"]\s*(?:javascript:|https?:|//)|<!DOCTYPE|<!ENTITY",
    re.IGNORECASE,
)


class ArtifactNotFoundError(FileNotFoundError):
    pass


def create_document_artifact(filename: str, content: str, format_type: str) -> dict[str, Any]:
    normalized_format = format_type.lower()
    if normalized_format not in {"txt", "md", "html", "json", "svg"}:
        raise ValueError("Document format must be txt, md, html, json, or svg")
    if len(content) > MAX_TEXT_CHARS:
        raise ValueError(f"Artifact content exceeds {MAX_TEXT_CHARS} characters")

    if normalized_format == "json":
        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Generated JSON is invalid: {exc.msg}") from exc
        content = json.dumps(parsed, ensure_ascii=False, indent=2)
    elif normalized_format == "svg":
        _validate_svg(content)

    extension, mime_type, kind = _FORMATS[normalized_format]
    safe_name = _safe_filename(filename, extension)
    return _write_artifact(
        filename=safe_name,
        data=content.encode("utf-8"),
        mime_type=mime_type,
        kind=kind,
    )


def create_table_artifact(
    filename: str,
    columns: list[str],
    rows: list[list[Any]],
    format_type: str,
    sheet_name: str = "AI Table",
) -> dict[str, Any]:
    if not columns:
        raise ValueError("A generated table requires at least one column")
    if len(columns) > 100:
        raise ValueError("A generated table cannot exceed 100 columns")
    if len(rows) > 10_000:
        raise ValueError("A generated table cannot exceed 10,000 rows")
    if any(len(row) != len(columns) for row in rows):
        raise ValueError("Every table row must have the same number of values as columns")

    normalized_format = format_type.lower()
    if normalized_format not in {"csv", "xlsx", "json"}:
        raise ValueError("Table format must be csv, xlsx, or json")

    if normalized_format == "csv":
        stream = io.StringIO(newline="")
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(_spreadsheet_safe_row(columns))
        writer.writerows([_spreadsheet_safe_row(row) for row in rows])
        data = ("\ufeff" + stream.getvalue()).encode("utf-8")
    elif normalized_format == "json":
        records = [dict(zip(columns, row)) for row in rows]
        data = json.dumps(records, ensure_ascii=False, indent=2, default=str).encode("utf-8")
    else:
        data = _build_xlsx(columns, rows, sheet_name)

    extension, mime_type, _ = _FORMATS[normalized_format]
    safe_name = _safe_filename(filename, extension)
    artifact = _write_artifact(
        filename=safe_name,
        data=data,
        mime_type=mime_type,
        kind="table",
    )
    artifact["row_count"] = len(rows)
    artifact["column_count"] = len(columns)
    _update_manifest(artifact)
    return artifact


def create_image_artifact(
    filename: str,
    data: bytes,
    mime_type: str,
    *,
    prompt: str | None = None,
) -> dict[str, Any]:
    normalized_mime = mime_type.split(";", 1)[0].strip().lower()
    extension = _IMAGE_MIME_TYPES.get(normalized_mime)
    if not extension:
        raise ValueError("Generated image must be PNG, JPEG, WebP, or GIF")
    _validate_raster_image(data, normalized_mime)
    artifact = _write_artifact(
        filename=_safe_filename(filename, extension),
        data=data,
        mime_type=normalized_mime,
        kind="image",
    )
    if prompt:
        artifact["prompt"] = prompt[:2000]
        _update_manifest(artifact)
    return artifact


def decode_image_data(value: str) -> tuple[bytes, str]:
    match = re.fullmatch(r"data:(image/(?:png|jpeg|webp|gif));base64,(.+)", value, re.DOTALL)
    if not match:
        raise ValueError("Image data must be a base64 PNG, JPEG, WebP, or GIF data URL")
    try:
        data = base64.b64decode(match.group(2), validate=True)
    except (ValueError, base64.binascii.Error) as exc:
        raise ValueError("Generated image contains invalid base64 data") from exc
    return data, match.group(1)


def get_artifact(artifact_id: str) -> tuple[dict[str, Any], Path]:
    if not _ARTIFACT_ID.fullmatch(artifact_id):
        raise ArtifactNotFoundError(artifact_id)
    directory = (ARTIFACT_DIR / artifact_id).resolve()
    manifest_path = directory / "manifest.json"
    if not manifest_path.is_file():
        raise ArtifactNotFoundError(artifact_id)
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        file_path = (directory / manifest["stored_name"]).resolve()
    except (OSError, KeyError, TypeError, json.JSONDecodeError) as exc:
        raise ArtifactNotFoundError(artifact_id) from exc
    if directory not in file_path.parents or not file_path.is_file():
        raise ArtifactNotFoundError(artifact_id)
    return _public_metadata(manifest), file_path


def _write_artifact(*, filename: str, data: bytes, mime_type: str, kind: str) -> dict[str, Any]:
    if not data:
        raise ValueError("Generated artifact cannot be empty")
    if len(data) > MAX_ARTIFACT_BYTES:
        raise ValueError(f"Generated artifact exceeds {MAX_ARTIFACT_BYTES} bytes")

    artifact_id = uuid.uuid4().hex
    directory = ARTIFACT_DIR / artifact_id
    directory.mkdir(parents=True, exist_ok=False)
    stored_name = f"artifact{Path(filename).suffix.lower()}"
    file_path = directory / stored_name
    file_path.write_bytes(data)

    manifest = {
        "artifact_id": artifact_id,
        "name": filename,
        "stored_name": stored_name,
        "kind": kind,
        "mime_type": mime_type,
        "size": len(data),
    }
    _write_manifest(directory, manifest)
    return _public_metadata(manifest)


def _update_manifest(artifact: dict[str, Any]) -> None:
    artifact_id = artifact["artifact_id"]
    directory = ARTIFACT_DIR / artifact_id
    manifest_path = directory / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    for key in ("row_count", "column_count", "prompt"):
        if key in artifact:
            manifest[key] = artifact[key]
    _write_manifest(directory, manifest)


def _write_manifest(directory: Path, manifest: dict[str, Any]) -> None:
    temp_path = directory / "manifest.json.tmp"
    temp_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(directory / "manifest.json")


def _public_metadata(manifest: dict[str, Any]) -> dict[str, Any]:
    artifact_id = manifest["artifact_id"]
    metadata = {
        key: manifest[key]
        for key in (
            "artifact_id",
            "name",
            "kind",
            "mime_type",
            "size",
            "row_count",
            "column_count",
            "prompt",
        )
        if key in manifest
    }
    metadata["preview_url"] = f"/ai/artifacts/{artifact_id}"
    metadata["download_url"] = f"/ai/artifacts/{artifact_id}/download"
    return metadata


def _safe_filename(filename: str, extension: str) -> str:
    name = Path(str(filename or "").replace("\\", "/")).name.strip()
    stem = Path(name).stem if name else "ai-artifact"
    stem = re.sub(r"[^\w .()-]", "_", stem, flags=re.UNICODE).strip(" .")
    if not stem:
        stem = "ai-artifact"
    return f"{stem[:180]}{extension}"


def _validate_svg(content: str) -> None:
    stripped = content.strip()
    if not stripped.lower().startswith("<svg") or "</svg>" not in stripped.lower():
        raise ValueError("Generated SVG must contain one complete <svg> element")
    if _UNSAFE_SVG.search(stripped):
        raise ValueError("Generated SVG contains unsafe active or external content")


def _validate_raster_image(data: bytes, mime_type: str) -> None:
    if not data or len(data) > MAX_ARTIFACT_BYTES:
        raise ValueError("Generated image is empty or too large")
    try:
        from PIL import Image

        with Image.open(io.BytesIO(data)) as image:
            image.verify()
            actual = (image.format or "").upper()
    except Exception as exc:
        raise ValueError("Generated image data is invalid") from exc
    expected = {
        "image/png": "PNG",
        "image/jpeg": "JPEG",
        "image/webp": "WEBP",
        "image/gif": "GIF",
    }[mime_type]
    if actual != expected:
        raise ValueError(f"Generated image data does not match {mime_type}")


def _build_xlsx(columns: list[str], rows: list[list[Any]], sheet_name: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - environment dependency guard
        raise RuntimeError("openpyxl is required to export generated XLSX tables") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = re.sub(r"[\\/*?:\[\]]", "_", sheet_name or "AI Table")[:31]
    sheet.append(_spreadsheet_safe_row(columns))
    for row in rows:
        sheet.append([_xlsx_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="EDE9FE")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="4C1D95")
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        sample_values = [str(column), *[str(row[index - 1] or "") for row in rows[:200]]]
        width = min(60, max(10, max(len(value) for value in sample_values) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _spreadsheet_safe_row(row: list[Any]) -> list[Any]:
    return [_xlsx_value(value) for value in row]
