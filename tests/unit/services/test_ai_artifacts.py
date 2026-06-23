import asyncio
import base64
import io
import json
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from PIL import Image

from services.ai_gateway import artifacts, image_generation
from services.ai_gateway import agent_handler
from services.ai_gateway.agent_handler import (
    AgentRequestPayload,
    AgentStep,
    _collect_generated_artifacts,
    _trace_tool_arguments,
    handle_agent,
)
from services.ai_gateway.function_registry import (
    AIFunctionInvokeRequest,
    get_registered_openai_tools,
    invoke_registered_function,
)


def _run(awaitable):
    return asyncio.run(awaitable)


@pytest.fixture(autouse=True)
def artifact_directory(monkeypatch, tmp_path):
    monkeypatch.setattr(artifacts, "ARTIFACT_DIR", tmp_path / "artifacts")


def test_generated_csv_table_is_persistent_and_exportable():
    metadata = artifacts.create_table_artifact(
        "results",
        ["class", "area_km2"],
        [["forest", 12.5], ["water", 3.25]],
        "csv",
    )

    loaded, path = artifacts.get_artifact(metadata["artifact_id"])

    assert loaded["kind"] == "table"
    assert loaded["name"] == "results.csv"
    assert loaded["row_count"] == 2
    assert loaded["download_url"].endswith("/download")
    assert path.read_text(encoding="utf-8-sig").splitlines() == [
        "class,area_km2",
        "forest,12.5",
        "water,3.25",
    ]


def test_generated_spreadsheets_neutralize_formula_injection():
    csv_metadata = artifacts.create_table_artifact(
        "safe.csv",
        ["value"],
        [["=HYPERLINK(\"https://example.test\")"]],
        "csv",
    )
    xlsx_metadata = artifacts.create_table_artifact(
        "safe.xlsx",
        ["value"],
        [["=HYPERLINK(\"https://example.test\")"]],
        "xlsx",
    )

    _, csv_path = artifacts.get_artifact(csv_metadata["artifact_id"])
    _, xlsx_path = artifacts.get_artifact(xlsx_metadata["artifact_id"])

    assert "'=HYPERLINK" in csv_path.read_text(encoding="utf-8-sig")
    assert load_workbook(xlsx_path)["AI Table"]["A2"].value.startswith("'=HYPERLINK")


def test_generated_xlsx_table_has_formatted_headers():
    metadata = artifacts.create_table_artifact(
        "analysis.xlsx",
        ["id", "label"],
        [[1, "healthy"], [2, "stressed"]],
        "xlsx",
        "Results",
    )
    _, path = artifacts.get_artifact(metadata["artifact_id"])

    workbook = load_workbook(path)
    sheet = workbook["Results"]

    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].value == "id"
    assert sheet["A1"].font.bold is True
    assert sheet["B3"].value == "stressed"


def test_safe_svg_is_exported_and_active_svg_is_rejected():
    metadata = artifacts.create_document_artifact(
        "diagram",
        '<svg xmlns="http://www.w3.org/2000/svg"><circle cx="5" cy="5" r="4"/></svg>',
        "svg",
    )

    assert metadata["kind"] == "image"
    assert metadata["name"] == "diagram.svg"
    with pytest.raises(ValueError, match="unsafe"):
        artifacts.create_document_artifact(
            "bad.svg",
            '<svg xmlns="http://www.w3.org/2000/svg"><script>alert(1)</script></svg>',
            "svg",
        )


def test_artifact_tools_are_registered_and_directly_invokable():
    tool_names = {
        tool["function"]["name"]
        for tool in get_registered_openai_tools()
    }
    assert {
        "create_generated_document",
        "create_generated_table",
        "generate_ai_image",
    }.issubset(tool_names)

    result = _run(
        invoke_registered_function(
            AIFunctionInvokeRequest(
                name="create_generated_table",
                arguments={
                    "filename": "summary.csv",
                    "columns": ["name", "value"],
                    "rows": [["NDVI", 0.62]],
                    "format": "csv",
                },
            ),
            db=object(),
            vector_db=object(),
        )
    )

    assert result["result"]["kind"] == "table"
    assert result["result"]["download_url"].endswith("/download")


def test_agent_response_collects_generated_artifacts():
    artifact = artifacts.create_document_artifact("report.md", "# Report", "md")
    steps = [
        AgentStep(
            step=1,
            type="tool",
            tool_call_id="call-1",
            name="create_generated_document",
            arguments={},
            status="success",
            result={"status": "success", "result": artifact},
        )
    ]

    assert _collect_generated_artifacts(steps) == [artifact]


def test_generated_content_is_summarized_in_agent_trace():
    document_trace = _trace_tool_arguments(
        "create_generated_document",
        {"filename": "report.md", "format": "md", "content": "x" * 20_000},
    )
    table_trace = _trace_tool_arguments(
        "create_generated_table",
        {
            "filename": "results.xlsx",
            "format": "xlsx",
            "columns": ["class", "pixels"],
            "rows": [["forest", 40], ["water", 10]],
        },
    )

    assert document_trace == {
        "filename": "report.md",
        "format": "md",
        "content_chars": 20_000,
    }
    assert table_trace["column_count"] == 2
    assert table_trace["row_count"] == 2
    assert "rows" not in table_trace


def test_agent_creates_and_returns_downloadable_table(monkeypatch):
    responses = [
        SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(
                        content="",
                        tool_calls=[
                            SimpleNamespace(
                                id="call-table",
                                type="function",
                                function=SimpleNamespace(
                                    name="create_generated_table",
                                    arguments=json.dumps(
                                        {
                                            "filename": "agent-results.xlsx",
                                            "columns": ["class", "pixels"],
                                            "rows": [["forest", 40], ["water", 10]],
                                            "format": "xlsx",
                                        }
                                    ),
                                ),
                            )
                        ],
                    )
                )
            ]
        ),
        SimpleNamespace(
            choices=[SimpleNamespace(message=SimpleNamespace(content="Table ready.", tool_calls=[]))]
        ),
    ]

    async def fake_acompletion(**kwargs):
        return responses.pop(0)

    async def empty_workspace_context(db, vector_db, limit):
        return ""

    monkeypatch.setattr(agent_handler, "acompletion", fake_acompletion)
    monkeypatch.setattr(agent_handler, "_build_workspace_context", empty_workspace_context)
    result = _run(
        handle_agent(
            AgentRequestPayload(
                user_prompt="Create an exportable class table.",
                language="en",
                tool_names=["create_generated_table"],
            ),
            db=object(),
            vector_db=object(),
            model_name="test-model",
        )
    )

    assert result["answer"] == "Table ready."
    assert result["used_tools"] == ["create_generated_table"]
    assert result["artifacts"][0]["name"] == "agent-results.xlsx"
    assert result["artifacts"][0]["download_url"].endswith("/download")


def test_provider_generated_image_is_persisted(monkeypatch):
    output = io.BytesIO()
    Image.new("RGB", (2, 2), color="purple").save(output, format="PNG")

    async def fake_image_generation(**kwargs):
        assert kwargs["model"] == "openai/test-image"
        return {"data": [{"b64_json": base64.b64encode(output.getvalue()).decode("ascii")}]}

    monkeypatch.setenv("AI_IMAGE_MODEL", "openai/test-image")
    monkeypatch.delenv("AI_API_KEY", raising=False)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    metadata = _run(
        image_generation.generate_ai_image(
            prompt="A purple square",
            filename="purple.png",
            image_generation_func=fake_image_generation,
        )
    )

    assert metadata["kind"] == "image"
    assert metadata["mime_type"] == "image/png"
    assert artifacts.get_artifact(metadata["artifact_id"])[1].read_bytes() == output.getvalue()
